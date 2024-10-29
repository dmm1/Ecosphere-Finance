import matplotlib
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
matplotlib.use('Agg') # Use a non-interactive backend
import base64
from io import BytesIO
import matplotlib.pyplot as plt
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from .forms import TransactionForm
from .models import Transaction, Category
from django.shortcuts import render
from django.db.models import Sum
from .models import Transaction
from .utils import update_recurring_expenses, get_upcoming_transactions
from django.utils import timezone
from datetime import timedelta, date
import datetime

@login_required
def visualize_expenses(request):
    transactions = Transaction.objects.filter(user=request.user, transaction_type='expense')
    categories = [t.category.name for t in transactions]
    amounts = [t.amount for t in transactions]

    # Create the plot
    plt.figure(figsize=(10, 5))
    plt.bar(categories, amounts)
    plt.xlabel('Category')
    plt.ylabel('Amount')
    plt.title('Expenses by Category')

    # Save it to a buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()

    # Encode the image to base64 so it can be rendered in the HTML
    image = base64.b64encode(image_png).decode('utf-8')

    # Pass the image to the template
    return render(request, 'finance/visualize.html', {'image': image})

@login_required
def transaction_list(request):
    transactions = Transaction.objects.filter(user=request.user)
    return render(request, 'finance/transaction_list.html', {'transactions': transactions})

@login_required
def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user  # Assign the current user to the transaction
            transaction.save()
            return redirect('transaction_list')
        else:
            print(form.errors)  # Debug: Print form errors to console
    else:
        form = TransactionForm()
    
    return render(request, 'finance/add_transaction.html', {'form': form})

@login_required
def delete_transaction(request, transaction_id):
    transaction = Transaction.objects.get(pk=transaction_id)
    transaction.delete()
    return redirect('transaction_list') # Redirect to the transaction list page

@login_required
def edit_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, pk=transaction_id)

    # Format amount for display
    formatted_amount = "{:.2f}".format(transaction.amount).replace('.', '.')

    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            return redirect('transaction_list')
    else:
        form = TransactionForm(instance=transaction)
        form.initial['amount'] = formatted_amount  # Set formatted amount to initial value for rendering

    # Format next_due_date for input
    next_due_date_formatted = transaction.next_due_date.strftime('%Y-%m-%d') if transaction.next_due_date else ''

    # Format transaction date for input if applicable
    transaction_date_formatted = transaction.date.strftime('%Y-%m-%d') if transaction.date else ''

    return render(request, 'finance/edit_transaction.html', {
        'form': form,
        'transaction': transaction,
        'next_due_date_formatted': next_due_date_formatted,
        'transaction_date_formatted': transaction_date_formatted,
        'formatted_amount': formatted_amount,
    })

@login_required
def home(request):
    # Update recurring transactions
    update_recurring_expenses()

    # Get recent transactions
    recent_transactions = Transaction.objects.filter(
        user=request.user,
        is_automated=False  # Only show manual transactions in recent list
    ).order_by('-date')[:5]

    # Calculate financial metrics
    income = Transaction.objects.filter(
        user=request.user, 
        transaction_type='income'
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    expenses = Transaction.objects.filter(
        user=request.user, 
        transaction_type='expense'
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    savings = income - expenses

    # Get top spending categories
    top_categories = (
        Transaction.objects.filter(
            user=request.user, 
            transaction_type='expense'
        )
        .values('category__name')
        .annotate(total_amount=Sum('amount'))
        .order_by('-total_amount')[:3]
    )

    # Get upcoming transactions for next 30 days
    upcoming_income, upcoming_expenses = get_upcoming_transactions(
        user=request.user,
        days_ahead=30
    )

    return render(request, 'finance/home.html', {
        'transactions': recent_transactions,
        'income': income,
        'expenses': expenses,
        'savings': savings,
        'top_categories': top_categories,
        'upcoming_income': upcoming_income,
        'upcoming_expenses': upcoming_expenses,
    })


@login_required
def export_transactions_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions_report.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setTitle('Transactions Report')

    transactions = Transaction.objects.filter(user=request.user)

    p.drawString(100, 750, 'Transactions Report')
    p.drawString(100, 730, '----------------------------------------')

    y = 710
    for transaction in transactions:
        p.drawString(100, y, f'{transaction.date} - {transaction.transaction_type} - {transaction.amount}')
        y -= 20

    p.showPage()
    p.save()

    return response

@login_required
def export_transactions_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transactions_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Transactions Report'

    # Define the header row
    headers = ['Date', 'Transaction Type', 'Amount', 'Category', 'Description']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add the data rows
    transactions = Transaction.objects.filter(user=request.user)
    for row_num, transaction in enumerate(transactions, 2):
        worksheet.cell(row=row_num, column=1).value = transaction.date.strftime('%Y-%m-%d')
        worksheet.cell(row=row_num, column=2).value = transaction.transaction_type
        worksheet.cell(row=row_num, column=3).value = transaction.amount
        worksheet.cell(row=row_num, column=4).value = transaction.category.name if transaction.category else ''
        worksheet.cell(row=row_num, column=5).value = transaction.description

    workbook.save(response)
    return response